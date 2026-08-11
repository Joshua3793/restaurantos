"""
Build the accountant's inventory valuation workbook from a count export.

  python3 scripts/build_count_workbook.py docs/audits/count-valuation-aug1.json out.xlsx

Values are FORMULAS (quantity x price), never Python-computed literals, so the
sheet recalculates if a price is corrected again. The counted quantity and the
price are kept in separate columns for exactly that reason.
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1]
OUT = sys.argv[2]

data = json.load(open(SRC))
rows = data["rows"]

# Independent arithmetic over the source data, stated in the workbook so the
# reader can confirm Excel's own recalculation agrees with it.
_at = sum(r["qtyBase"] * r["priceAtCountPerBase"] for r in rows)
_now = sum(r["qtyBase"] * r["priceNowPerBase"] for r in rows)

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=14, bold=True)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="555555")
MONEY = '$#,##0.00;($#,##0.00);-'
MONEY5 = '$#,##0.00000;($#,##0.00000);-'
QTY = '#,##0.###'
THIN = Side(style="thin", color="D4D4D8")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
TOPLINE = Border(top=Side(style="medium", color="1F2937"))

wb = Workbook()

# ─────────────────────────────────────────────────────────── Valuation ──
ws = wb.active
ws.title = "Valuation"

ws["A1"] = f"Inventory valuation — {data['revenueCenter']} — {data['sessionDate']}"
ws["A1"].font = TITLE_FONT
ws["A2"] = (
    f"Source count: \"{data['label']}\" (physical count finalized "
    f"{(data['finalizedAt'] or '')[:10]}). Quantities are as counted; prices are the "
    f"corrected prices in the system as at {data['exportedAt'][:10]}."
)
ws["A2"].font = NOTE

HEADERS = [
    ("Item", 38), ("Category", 10), ("Storage area", 16), ("Supplier", 16),
    ("Quantity basis", 30), ("Counted on", 12), ("Qty counted", 12), ("Count unit", 12),
    ("Qty (base)", 12), ("Base unit", 10),
    ("Price at count ($/base)", 20), ("Price now ($/base)", 18),
    ("Value at count", 15), ("Value at current prices", 21), ("Change", 12),
    ("Price now (readable)", 19), ("Per", 8), ("Pack format", 30),
]
HROW = 4
for i, (h, w) in enumerate(HEADERS, start=1):
    c = ws.cell(row=HROW, column=i, value=h)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[HROW].height = 30
ws.freeze_panes = f"A{HROW + 1}"

first = HROW + 1
for n, r in enumerate(rows):
    i = first + n
    vals = [
        r["item"], r["category"], r["storageArea"], r["supplier"], r["basis"],
        r.get("countedOn", ""), r["countQty"], r["countUom"], r["qtyBase"], r["baseUnit"],
        r["priceAtCountPerBase"], r["priceNowPerBase"],
        f"=I{i}*K{i}", f"=I{i}*L{i}", f"=N{i}-M{i}",
        r["priceNowPerUnit"], r["perUnit"], r["packChain"],
    ]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = BODY
        c.border = BOX
    for col in (7, 9):
        ws.cell(row=i, column=col).number_format = QTY
    for col in (11, 12):
        ws.cell(row=i, column=col).number_format = MONEY5
    for col in (13, 14, 15, 16):
        ws.cell(row=i, column=col).number_format = MONEY

last = first + len(rows) - 1
tot = last + 1
ws.cell(row=tot, column=1, value="TOTAL").font = BOLD
for col in (13, 14, 15):
    L = get_column_letter(col)
    c = ws.cell(row=tot, column=col, value=f"=SUM({L}{first}:{L}{last})")
    c.font = BOLD
    c.number_format = MONEY
    c.border = TOPLINE
ws.cell(row=tot, column=1).border = TOPLINE

ws.auto_filter.ref = f"A{HROW}:R{last}"

# ───────────────────────────────────────────────────────────── Summary ──
s = wb.create_sheet("Summary")
s["A1"] = f"Summary — {data['revenueCenter']} — {data['sessionDate']}"
s["A1"].font = TITLE_FONT

s["A3"] = "By category"
s["A3"].font = BOLD
for i, h in enumerate(["Category", "Lines", "Value at count", "Value at current prices", "Change"], start=1):
    c = s.cell(row=4, column=i, value=h)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
cats = sorted({r["category"] for r in rows})
for n, cat in enumerate(cats):
    i = 5 + n
    s.cell(row=i, column=1, value=cat).font = BODY
    s.cell(row=i, column=2, value=f'=COUNTIF(Valuation!$B${first}:$B${last},$A{i})').font = BODY
    s.cell(row=i, column=3, value=f'=SUMIF(Valuation!$B${first}:$B${last},$A{i},Valuation!$M${first}:$M${last})').font = BODY
    s.cell(row=i, column=4, value=f'=SUMIF(Valuation!$B${first}:$B${last},$A{i},Valuation!$N${first}:$N${last})').font = BODY
    s.cell(row=i, column=5, value=f"=D{i}-C{i}").font = BODY
    for col in (3, 4, 5):
        s.cell(row=i, column=col).number_format = MONEY
crow = 5 + len(cats)
s.cell(row=crow, column=1, value="TOTAL").font = BOLD
for col in (2, 3, 4, 5):
    L = get_column_letter(col)
    c = s.cell(row=crow, column=col, value=f"=SUM({L}5:{L}{crow - 1})")
    c.font = BOLD
    c.border = TOPLINE
    if col > 2:
        c.number_format = MONEY
s.cell(row=crow, column=1).border = TOPLINE

brow = crow + 3
s.cell(row=brow, column=1, value="How each quantity was established").font = BOLD
s.cell(row=brow + 1, column=1, value=(
    "A full count values every line, including lines nobody counted. Lines marked "
    "\"carried\" or \"not counted\" take the quantity the system expected rather than a "
    "physical observation."
)).font = NOTE
for i, h in enumerate(["Quantity basis", "Lines", "Value at current prices", "% of total"], start=1):
    c = s.cell(row=brow + 2, column=i, value=h)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
# Derived from the data — a multi-session month-end produces bases like
# "Physically counted 2026-07-31", which no hardcoded list would anticipate.
bases = sorted({r["basis"] for r in rows},
               key=lambda b: (not b.startswith("Physically"), b))
for n, b in enumerate(bases):
    i = brow + 3 + n
    s.cell(row=i, column=1, value=b).font = BODY
    s.cell(row=i, column=2, value=f'=COUNTIF(Valuation!$E${first}:$E${last},$A{i})').font = BODY
    s.cell(row=i, column=3, value=f'=SUMIF(Valuation!$E${first}:$E${last},$A{i},Valuation!$N${first}:$N${last})').font = BODY
    s.cell(row=i, column=3).number_format = MONEY
    s.cell(row=i, column=4, value=f"=IF($C${brow + 3 + len(bases)}=0,0,C{i}/$C${brow + 3 + len(bases)})").font = BODY
    s.cell(row=i, column=4).number_format = "0.0%"
trow = brow + 3 + len(bases)
s.cell(row=trow, column=1, value="TOTAL").font = BOLD
for col in (2, 3):
    L = get_column_letter(col)
    c = s.cell(row=trow, column=col, value=f"=SUM({L}{brow + 3}:{L}{trow - 1})")
    c.font = BOLD
    c.border = TOPLINE
    if col == 3:
        c.number_format = MONEY
s.cell(row=trow, column=1).border = TOPLINE

for col, w in zip("ABCDE", (34, 10, 22, 24, 14)):
    s.column_dimensions[col].width = w

# ─────────────────────────────────────────────────────────────── Notes ──
n = wb.create_sheet("Notes")
n["A1"] = "Basis of preparation"
n["A1"].font = TITLE_FONT
n.column_dimensions["A"].width = 118

LINES = [
    "",
    f"Scope. Revenue centre {data['revenueCenter']}, which is the default revenue centre — its stock is held on the "
    f"inventory record itself rather than as a per-centre allocation. {len(rows)} items, being every item that appeared "
    "on either count sheet.",
    "",
    "Sources. The month-end was counted across more than one session: " + data["label"] + ". No single session holds "
    "the whole position — items counted in one session appear in the other as carried-forward lines. Each item is "
    "therefore taken from the session in which it was PHYSICALLY counted, shown per line in 'Quantity basis' and "
    "'Counted on'.",
    "",
    f"Quantities. As recorded by the count itself, not recomputed. Several items' pack formats have since been "
    "corrected, and re-resolving a counter's entry through a corrected pack would answer a different question than the "
    "one they answered on the day — so the quantity each count froze is used unchanged. Quantities are shown "
    "both in the unit the counter worked in (Qty counted / Count unit) and in the item's base unit, which is what the "
    "system stores and what the valuation multiplies.",
    "",
    "Prices. The count froze a price against each line when it was finalized. Several of those prices were "
    "subsequently found to be wrong: when a supplier changed pack size, the new case price was divided by the item's "
    "old pack size, so the cost was wrong by the ratio between them. Those items have since been corrected. This "
    "workbook therefore values the counted quantities at the CORRECTED prices, and shows the price used at count time "
    "alongside, so the restatement of each line is visible.",
    "",
    "Both figures are given per line: 'Value at count' reproduces the original count total, and 'Value at current "
    "prices' is the restated figure. The difference is in the Change column and is subtotalled by category on the "
    "Summary sheet.",
    "",
    "Quantity basis — please read before signing. A full count in this system values every line on the sheet, "
    "including lines that were never physically counted. Where no quantity was entered, the line is valued at the "
    "quantity the system expected. The Summary sheet splits the total on that basis. Only the 'Physically counted' "
    "portion rests on an observation; the remainder is a carry-forward from the previous count adjusted for recorded "
    "movements.",
    "",
    "Items subsequently deactivated remain in this valuation, because they were on hand at the count date.",
    "",
    "Formulas. Value columns are formulas over the quantity and price columns, so correcting a price in this workbook "
    "updates the totals. Nothing is a pasted result.",
    "",
    "Check figures. The value columns are live formulas, so Excel computes them when this file opens. They should "
    f"total ${_at:,.2f} at count prices and ${_now:,.2f} at current prices, a change of ${_now - _at:,.2f}. The first "
    f"of those reproduces the total the system recorded when the count was finalized (${data['storedTotal']:,.2f}), "
    "which is the check that these quantities are the count's own. If the totals differ from these figures, a price "
    "has been edited since this file was produced.",
    "",
    "Source count sessions: " + ", ".join(data["sessionIds"]),
    f"Count finalized: {data['finalizedAt'] or 'n/a'}",
    f"Prices as at: {data['exportedAt'][:19].replace('T', ' ')} UTC",
    f"Count total as stored by the system at finalize: ${data['storedTotal']:,.2f}",
]
for i, t in enumerate(LINES, start=2):
    c = n.cell(row=i, column=1, value=t)
    c.font = BODY if t and not t.startswith(("Source count session", "Count finalized", "Prices as at", "Count total")) else NOTE
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if t:
        n.row_dimensions[i].height = max(14, 13 * (len(t) // 115 + 1))

wb.save(OUT)
print(f"wrote {OUT}: {len(rows)} lines, sheets {wb.sheetnames}")
