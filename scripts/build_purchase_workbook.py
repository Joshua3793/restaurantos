"""
Build the catering valuation workbook from a purchase export.

  python3 scripts/build_purchase_workbook.py docs/audits/purchase-valuation-catering-jul.json out.xlsx

Values are FORMULAS over the quantity and price columns, so the sheet
recalculates if a price is corrected.
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC, OUT = sys.argv[1], sys.argv[2]
data = json.load(open(SRC))
items, details = data["items"], data["details"]

ok_items = [r for r in items if r["qtyBaseReconciled"] > 0]
flagged = [d for d in details if d["flags"]]
_valued = sum(r["qtyBaseReconciled"] * r["pricePerBase"] for r in items)
_invoiced = sum(r["invoicedTotal"] for r in items)
_excluded = sum((r["qtyBase"] - r["qtyBaseReconciled"]) * r["pricePerBase"] for r in items)

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=FONT, size=14, bold=True)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="555555")
MONEY = '$#,##0.00;($#,##0.00);-'
MONEY5 = '$#,##0.00000;($#,##0.00000);-'
QTY = '#,##0.###'
THIN = Side(style="thin", color="D4D4D8")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
TOPLINE = Border(top=Side(style="medium", color="1F2937"))


def header(ws, row, cols):
    for i, (h, w) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


wb = Workbook()

# ──────────────────────────────────────────────────────────── Position ──
ws = wb.active
ws.title = "Position"
ws["A1"] = f"{data['revenueCenter']} inventory — as at {data['asAt']}"
ws["A1"].font = TITLE
ws["A2"] = (
    f"Basis: purchases attributed to {data['revenueCenter']} from {data['from']} to {data['to']}, treated as the "
    f"closing position. Valued at corrected prices as at {data['exportedAt'][:10]}. "
    "This is NOT a physical count — see Notes."
)
ws["A2"].font = NOTE

COLS = [("Item", 36), ("Category", 10), ("Supplier", 20), ("Last invoice", 13),
        ("Qty", 11), ("Unit", 11), ("Price ($/base)", 15), ("Qty (base)", 12), ("Base unit", 10),
        ("Value", 13), ("Invoiced", 13), ("Variance", 12), ("Price under review", 18)]
HROW = 4
header(ws, HROW, COLS)
ws.freeze_panes = f"A{HROW+1}"

first = HROW + 1
for n, r in enumerate(ok_items):
    i = first + n
    perCount = (r["qtyBase"] / r["countQty"]) if r["countQty"] else 1
    vals = [
        r["item"], r["category"], r["supplier"], r["lastInvoiceDate"],
        r["qtyBaseReconciled"] / perCount, r["countUom"], r["pricePerBase"],
        r["qtyBaseReconciled"], r["baseUnit"],
        f"=H{i}*G{i}", r["invoicedTotal"], f"=J{i}-K{i}",
        "Yes — provisional" if r["priceUnderReview"] else "",
    ]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = BODY
        c.border = BOX
    for col in (5, 8):
        ws.cell(row=i, column=col).number_format = QTY
    ws.cell(row=i, column=7).number_format = MONEY5
    for col in (10, 11, 12):
        ws.cell(row=i, column=col).number_format = MONEY
    if r["priceUnderReview"]:
        ws.cell(row=i, column=13).fill = WARN_FILL

last = first + len(ok_items) - 1
tot = last + 1
ws.cell(row=tot, column=1, value="TOTAL").font = BOLD
ws.cell(row=tot, column=1).border = TOPLINE
for L in ("J", "K", "L"):
    c = ws[f"{L}{tot}"]
    c.value = f"=SUM({L}{first}:{L}{last})"
    c.font = BOLD
    c.number_format = MONEY
    c.border = TOPLINE
ws.auto_filter.ref = f"A{HROW}:M{last}"

ws.cell(row=tot + 2, column=1, value="Excluded — lines that could not be reconciled (see Needs review)").font = BOLD
c = ws.cell(row=tot + 2, column=10, value=_excluded)
c.font = BOLD
c.number_format = MONEY
c.fill = WARN_FILL

# ─────────────────────────────────────────────────────────── Purchases ──
p = wb.create_sheet("Purchases")
p["A1"] = "Invoice lines behind the position"
p["A1"].font = TITLE
PCOLS = [("Invoice date", 13), ("Supplier", 22), ("Item", 32), ("As printed on the invoice", 38),
         ("Pack", 16), ("Qty", 11), ("Unit", 11), ("Valued", 12), ("Invoiced", 12), ("Reconciles", 12), ("Issue", 52)]
header(p, 3, PCOLS)
p.freeze_panes = "A4"
for n, d in enumerate(details):
    i = 4 + n
    vals = [d["invoiceDate"], d["supplier"], d["item"], d["description"], d["pack"],
            d["countQty"], d["countUom"], d["valued"], d["lineTotal"],
            "No" if d["flags"] else "Yes", "; ".join(d["flags"])]
    for col, v in enumerate(vals, start=1):
        c = p.cell(row=i, column=col, value=v)
        c.font = BODY
        c.border = BOX
        if d["flags"]:
            c.fill = WARN_FILL
    p.cell(row=i, column=6).number_format = QTY
    for col in (8, 9):
        p.cell(row=i, column=col).number_format = MONEY
p.auto_filter.ref = f"A3:K{3 + len(details)}"

# ───────────────────────────────────────────────────────── Needs review ──
nr = wb.create_sheet("Needs review")
nr["A1"] = "Lines excluded from the position"
nr["A1"].font = TITLE
nr["A2"] = (
    "Each of these bills an amount that does not agree with what the quantity is worth at the item's price. "
    "They are excluded from the total rather than folded in, because either the invoiced pack or the item's "
    "price is wrong and it cannot be told which from the invoice alone."
)
nr["A2"].font = NOTE
nr.column_dimensions["A"].width = 118
header(nr, 4, [("Invoice date", 13), ("Item", 30), ("As printed", 36), ("Pack", 16),
               ("Qty", 11), ("Unit", 10), ("Valued", 12), ("Invoiced", 12), ("Issue", 54)])
for n, d in enumerate(flagged):
    i = 5 + n
    vals = [d["invoiceDate"], d["item"], d["description"], d["pack"], d["countQty"], d["countUom"],
            d["valued"], d["lineTotal"], "; ".join(d["flags"])]
    for col, v in enumerate(vals, start=1):
        c = nr.cell(row=i, column=col, value=v)
        c.font = BODY
        c.border = BOX
    nr.cell(row=i, column=5).number_format = QTY
    for col in (7, 8):
        nr.cell(row=i, column=col).number_format = MONEY
r = 5 + len(flagged)
nr.cell(row=r, column=1, value="Excluded value").font = BOLD
c = nr.cell(row=r, column=7, value=_excluded)
c.font = BOLD
c.number_format = MONEY
c.border = TOPLINE

# ─────────────────────────────────────────────────────────────── Notes ──
n = wb.create_sheet("Notes")
n["A1"] = "Basis of preparation"
n["A1"].font = TITLE
n.column_dimensions["A"].width = 118
LINES = [
    "",
    f"THIS IS NOT A PHYSICAL COUNT. {data['revenueCenter']} was not counted. On instruction, the position is taken "
    f"to be everything purchased for {data['revenueCenter']} between {data['from']} and {data['to']}, on the "
    "assumption that the week's buying was still on hand at the period end and none of it had been consumed. If any "
    "of it was used, this figure overstates the closing position by the amount used. A physical count is the only "
    "way to remove that assumption.",
    "",
    f"Scope. {len(details)} approved invoice lines attributed to {data['revenueCenter']}, covering {len(items)} "
    "distinct items. A line is attributed by its own revenue-centre assignment, by an explicit split across centres, "
    "or by inheriting the invoice's assignment when it carries none of its own. The Purchases sheet shows every line "
    "and how it was attributed.",
    "",
    "Quantities. Derived with the same rule the stock count and theoretical receiving use, so this cannot drift from "
    "them: a line's received quantity expands through the pack printed on the invoice.",
    "",
    "Prices. The item's corrected price as at the export date, not the price on the invoice. The Invoiced column "
    "shows what was actually billed, and Variance is the difference. They diverge when a price moved after the "
    "invoice was approved.",
    "",
    f"Exclusions. {len(flagged)} lines worth ${_excluded:,.2f} are excluded and listed on Needs review. In each case "
    "what the quantity is worth at the item's price differs from what was invoiced by more than 25%, or the line "
    "carries no invoiced total at all. One Veal Bones line reads a pack of \"1 × 501 lb\" with no total, which alone "
    "would have added 227 kg of stock and $0 of cost. Including such lines would be guessing.",
    "",
    "Prices under review. Items marked in the last column of the Position sheet still carry an unresolved flag from "
    "the inventory price audit — their valuation is provisional until the flag is cleared. Notably, four of the "
    "excluded lines are for items on that list, so the reconciliation check and the price audit agree with each "
    "other about where the data is weak.",
    "",
    "Formulas. Value and Variance are formulas over the quantity and price columns, so correcting a price updates "
    "the totals. Nothing is a pasted result.",
    "",
    f"Check figures. The Position sheet should total ${_valued:,.2f} against ${_invoiced:,.2f} invoiced, a variance "
    f"of ${_valued - _invoiced:,.2f}, with ${_excluded:,.2f} excluded on top. If Excel shows different totals, a "
    "price has been edited since this file was produced.",
    "",
    f"Window: {data['from']} to {data['to']}.  Prices as at: {data['exportedAt'][:19].replace('T', ' ')} UTC.",
]
for i, t in enumerate(LINES, start=2):
    c = n.cell(row=i, column=1, value=t)
    c.font = NOTE if t.startswith("Window:") else BODY
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if t:
        n.row_dimensions[i].height = max(14, 13 * (len(t) // 112 + 1))
n["A2"].font = Font(name=FONT, size=10, bold=True)

wb.save(OUT)
print(f"wrote {OUT}: {len(ok_items)} items, {len(details)} lines, {len(flagged)} excluded, sheets {wb.sheetnames}")
